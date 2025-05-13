# sports_clubs.py

from flask import (
    Blueprint, render_template, request, flash, redirect, url_for,
    send_file, 
    current_app,
    session 
)
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
import os
from datetime import datetime, date
from io import BytesIO
import traceback
import uuid 
import shutil 
import openpyxl 
from openpyxl.styles import Font, numbers 
import re # Added for sanitizing sheet names
from decimal import Decimal # Added for financial calculations

from flask_wtf import FlaskForm
from wtforms import SubmitField, FileField, SelectField, DateField, StringField, TextAreaField 
from wtforms.validators import DataRequired, Optional, Regexp, NumberRange, ValidationError
from wtforms.fields import DecimalField as WTDecimalField

# Import necessary functions and constants from processing_utils
from processing_utils import (
    update_summary_file, 
    CLUB_SHEET_HEADERS, 
    NEEDS_REVIEW_SHEET_NAME, 
    SUMMARY_SHEET_NAME, 
    SUMMARY_INDIVIDUAL_SHEET_NAME,
    SUMMARY_CLUB_COL,
    SUMMARY_ROLLOVER_COL,
    SUMMARY_CONTRIB_COL,
    SUMMARY_CHARGES_COL,
    SUMMARY_EXPENSES_COL,
    SUMMARY_REMAINING_COL,
    to_decimal # Assuming this helper is accessible
)

try:
    from admin import admin_required 
except ImportError:
    from functools import wraps
    def admin_required(f): 
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated or getattr(current_user, 'role', None) != 'admin':
                flash('Admin access required.', 'danger')
                return redirect(url_for('auth.login')) 
            return f(*args, **kwargs)
        return decorated_function

from models import FoundationTransaction, Club 
from db import SessionLocal 

ALLOWED_EXTENSIONS = {'xlsx'}

sc_bp = Blueprint('sports_clubs', __name__, url_prefix='/admin/sports-clubs', template_folder='templates')

def get_temp_file_storage_path(create=True):
    path = os.path.join(current_app.instance_path, 'temp_summary_files')
    if create and not os.path.exists(path):
        try:
            os.makedirs(path)
            current_app.logger.info(f"Created temporary file directory: {path}")
        except OSError as e:
            current_app.logger.error(f"Could not create temp_summary_files directory: {path}. Error: {e}")
            return None 
    return path

class FoundationReconciliationForm(FlaskForm):
    drs_report = FileField('DRS FS Report (Excel)', validators=[DataRequired()])
    donor_report = FileField('Donor Report (Excel)', validators=[DataRequired()])
    summary_file = FileField('Current Foundation Summary (Excel)', validators=[DataRequired()])
    submit = SubmitField('Process Files')

class ManualTransactionForm(FlaskForm):
    transaction_type = SelectField('Transaction Type', choices=[
        ('Contribution', 'Contribution'),
        ('Expense', 'Expense/Fee'), # Represents a debit from club funds
        ('Refund', 'Refund (to club)') # Represents a credit to club funds, not from donation
    ], validators=[DataRequired()])
    transaction_date = DateField('Transaction Date', validators=[DataRequired()], default=date.today)
    journal_ref = StringField('Journal Reference (if any)', validators=[Optional()])
    name_description = StringField('Name / Description', validators=[DataRequired(message="Please provide a name or description.")])
    amount = WTDecimalField('Amount (Positive Value)', validators=[DataRequired(), NumberRange(min=0.01)], places=2)
    club_name = SelectField('Assign to Club', validators=[DataRequired(message="Please select a club.")])
    notes = TextAreaField('Notes (Optional)', validators=[Optional()])
    submit = SubmitField('Add Manual Entry')

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# --- Helper Function to Update Excel After Review ---
def _update_temp_summary_after_review(reviewed_transaction: FoundationTransaction, assigned_club: Club):
    """
    Loads the existing temporary summary Excel, moves the reviewed_transaction
    from 'Needs Review' to the assigned_club's sheet, recalculates summaries,
    and saves the Excel back to the temporary file.
    Returns True on success, False on failure.
    """
    temp_filename = session.get('last_summary_temp_filename')
    temp_file_storage_path = get_temp_file_storage_path(create=False)

    if not temp_filename or not temp_file_storage_path:
        current_app.logger.error("Cannot update temp summary: No temp file info in session or path not found.")
        return False

    temp_filepath = os.path.join(temp_file_storage_path, temp_filename)
    if not os.path.exists(temp_filepath):
        current_app.logger.error(f"Cannot update temp summary: Temp file {temp_filepath} does not exist.")
        return False

    try:
        workbook = openpyxl.load_workbook(temp_filepath)
        current_app.logger.info(f"Loaded temp summary {temp_filepath} for update after review.")

        # --- 1. Find and Remove from 'Needs Review' sheet ---
        needs_review_ws = None
        if NEEDS_REVIEW_SHEET_NAME in workbook.sheetnames:
            needs_review_ws = workbook[NEEDS_REVIEW_SHEET_NAME] # Corrected: Use dictionary-style access
        
        if needs_review_ws:
            row_to_delete = None
            jrnl_ref_col_idx_nr = -1
            # Assuming CLUB_SHEET_HEADERS[2] is 'Journal Ref'
            for c_idx, cell in enumerate(needs_review_ws[1]): 
                if cell.value == CLUB_SHEET_HEADERS[2]: 
                    jrnl_ref_col_idx_nr = c_idx + 1
                    break
            if jrnl_ref_col_idx_nr != -1:
                for r_idx in range(needs_review_ws.max_row, 1, -1): 
                    if needs_review_ws.cell(row=r_idx, column=jrnl_ref_col_idx_nr).value == reviewed_transaction.journal_ref:
                        row_to_delete = r_idx
                        break
            if row_to_delete:
                needs_review_ws.delete_rows(row_to_delete)
                current_app.logger.info(f"Removed transaction {reviewed_transaction.journal_ref} from '{NEEDS_REVIEW_SHEET_NAME}'.")
            else:
                current_app.logger.warning(f"Transaction {reviewed_transaction.journal_ref} not found in '{NEEDS_REVIEW_SHEET_NAME}'.")
        else:
            current_app.logger.warning(f"'{NEEDS_REVIEW_SHEET_NAME}' sheet not found in workbook.")

        # --- 2. Add to Assigned Club's Sheet ---
        safe_club_sheet_name = re.sub(r'[\\/*?:\[\]]', '_', assigned_club.name)[:31]
        if safe_club_sheet_name.lower().endswith(" club"):
            safe_club_sheet_name = safe_club_sheet_name[:-5].strip()
        
        club_ws = None
        if safe_club_sheet_name in workbook.sheetnames:
            club_ws = workbook[safe_club_sheet_name] # Corrected
        
        if not club_ws: 
            current_app.logger.info(f"Club sheet '{safe_club_sheet_name}' not found, creating it.")
            club_ws = workbook.create_sheet(title=safe_club_sheet_name)
            club_ws.append(CLUB_SHEET_HEADERS)
            for cell_header in club_ws[1]: cell_header.font = Font(bold=True)
            club_ws.freeze_panes = 'A2'

        row_data = [
            reviewed_transaction.transaction_date,
            "Contribution" if reviewed_transaction.gross_amount > 0 else "Fee/Expense",
            reviewed_transaction.journal_ref,
            reviewed_transaction.donor_description,
            reviewed_transaction.gross_amount,
            reviewed_transaction.fees_total,
            reviewed_transaction.net_amount,
            assigned_club.name
        ]
        club_ws.append(row_data)
        new_row_idx = club_ws.max_row
        club_ws.cell(row=new_row_idx, column=1).number_format = numbers.FORMAT_DATE_YYYYMMDD2
        for col_idx_currency in [5, 6, 7]: # Contribution Amount, Charges/Offset, Net Amount
            club_ws.cell(row=new_row_idx, column=col_idx_currency).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        current_app.logger.info(f"Added transaction {reviewed_transaction.journal_ref} to '{safe_club_sheet_name}'.")

        # --- 3. Update 'Summary Individual' sheet ---
        summary_individual_ws = None
        if SUMMARY_INDIVIDUAL_SHEET_NAME in workbook.sheetnames:
            summary_individual_ws = workbook[SUMMARY_INDIVIDUAL_SHEET_NAME] # Corrected

        if summary_individual_ws:
            jrnl_ref_col_idx_si = -1; club_assign_col_idx_si = -1
            for c_idx, cell in enumerate(summary_individual_ws[1]):
                if cell.value == CLUB_SHEET_HEADERS[2]: jrnl_ref_col_idx_si = c_idx + 1
                if cell.value == CLUB_SHEET_HEADERS[7]: club_assign_col_idx_si = c_idx + 1
            
            updated_in_summary_individual = False
            if jrnl_ref_col_idx_si != -1 and club_assign_col_idx_si != -1:
                for r_idx in range(2, summary_individual_ws.max_row + 1):
                     if summary_individual_ws.cell(row=r_idx, column=jrnl_ref_col_idx_si).value == reviewed_transaction.journal_ref:
                        summary_individual_ws.cell(row=r_idx, column=club_assign_col_idx_si).value = assigned_club.name
                        updated_in_summary_individual = True
                        current_app.logger.info(f"Updated club assignment for {reviewed_transaction.journal_ref} in '{SUMMARY_INDIVIDUAL_SHEET_NAME}'.")
                        break
            if not updated_in_summary_individual: 
                summary_individual_ws.append(row_data) 
                new_si_row_idx = summary_individual_ws.max_row
                summary_individual_ws.cell(row=new_si_row_idx, column=1).number_format = numbers.FORMAT_DATE_YYYYMMDD2
                for col_idx_currency_si in [5, 6, 7]:
                    summary_individual_ws.cell(row=new_si_row_idx, column=col_idx_currency_si).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                current_app.logger.info(f"Appended transaction {reviewed_transaction.journal_ref} to '{SUMMARY_INDIVIDUAL_SHEET_NAME}'.")
        else:
            current_app.logger.warning(f"'{SUMMARY_INDIVIDUAL_SHEET_NAME}' sheet not found.")


        # --- 4. Recalculate 'Summary' Sheet ---
        current_app.logger.info("Recalculating main 'Summary' sheet...")
        ws_summary = None
        if SUMMARY_SHEET_NAME in workbook.sheetnames:
            ws_summary = workbook[SUMMARY_SHEET_NAME] # Corrected
        
        if not ws_summary:
            current_app.logger.error(f"'{SUMMARY_SHEET_NAME}' sheet not found for recalculation.")
            return False

        summary_headers_map = {}
        summary_header_row_idx = -1
        
        required_summary_headers_list = [SUMMARY_CLUB_COL, SUMMARY_ROLLOVER_COL, SUMMARY_CONTRIB_COL, SUMMARY_CHARGES_COL, SUMMARY_EXPENSES_COL, SUMMARY_REMAINING_COL]
        for r_idx in range(1, ws_summary.max_row + 1):
            row_vals = [str(c.value).strip() if c.value is not None else "" for c in ws_summary[r_idx]]
            if all(req_h in row_vals for req_h in required_summary_headers_list):
                summary_header_row_idx = r_idx
                for c_idx, header_val in enumerate(row_vals):
                    if header_val: summary_headers_map[header_val] = c_idx + 1
                break
        
        if summary_header_row_idx == -1 or not all(h in summary_headers_map for h in required_summary_headers_list):
            current_app.logger.error(f"Could not find all required headers in '{SUMMARY_SHEET_NAME}' for recalculation.")
            return False

        col_idx_s_club = summary_headers_map[SUMMARY_CLUB_COL]
        col_idx_s_rollover = summary_headers_map[SUMMARY_ROLLOVER_COL]
        col_idx_s_contrib = summary_headers_map[SUMMARY_CONTRIB_COL]
        col_idx_s_charges = summary_headers_map[SUMMARY_CHARGES_COL]
        col_idx_s_expenses = summary_headers_map[SUMMARY_EXPENSES_COL]
        col_idx_s_remaining = summary_headers_map[SUMMARY_REMAINING_COL]

        today_date = date.today()
        current_fy_start_date = date(today_date.year if today_date.month >= 7 else today_date.year - 1, 7, 1)

        for r_idx in range(summary_header_row_idx + 1, ws_summary.max_row + 1):
            club_name_summary = ws_summary.cell(row=r_idx, column=col_idx_s_club).value
            if not club_name_summary or str(club_name_summary).strip().lower() == 'grand total':
                continue
            
            club_name_summary_str = str(club_name_summary).strip()
            safe_club_sheet_name_lookup = re.sub(r'[\\/*?:\[\]]', '_', club_name_summary_str)[:31]

            fy_contributions = Decimal('0.00')
            fy_charges_offsets = Decimal('0.00')

            if safe_club_sheet_name_lookup in workbook.sheetnames:
                club_data_ws = workbook[safe_club_sheet_name_lookup] # Corrected
                club_sheet_headers_map = {}
                if club_data_ws.max_row > 0: 
                    for c_idx_club, cell_club_h in enumerate(club_data_ws[1]): 
                        if cell_club_h.value: club_sheet_headers_map[str(cell_club_h.value).strip()] = c_idx_club + 1
                
                date_col_club = club_sheet_headers_map.get(CLUB_SHEET_HEADERS[0]) 
                contrib_col_club = club_sheet_headers_map.get(CLUB_SHEET_HEADERS[4]) 
                charges_col_club = club_sheet_headers_map.get(CLUB_SHEET_HEADERS[5]) 

                if date_col_club and contrib_col_club and charges_col_club:
                    for club_r_idx in range(2, club_data_ws.max_row + 1): 
                        transaction_date_val = club_data_ws.cell(row=club_r_idx, column=date_col_club).value
                        transaction_date_obj = None
                        if isinstance(transaction_date_val, datetime): transaction_date_obj = transaction_date_val.date()
                        elif isinstance(transaction_date_val, date): transaction_date_obj = transaction_date_val
                        
                        if transaction_date_obj and transaction_date_obj >= current_fy_start_date:
                            fy_contributions += to_decimal(club_data_ws.cell(row=club_r_idx, column=contrib_col_club).value)
                            fy_charges_offsets += to_decimal(club_data_ws.cell(row=club_r_idx, column=charges_col_club).value)
            
            rollover_val_s = to_decimal(ws_summary.cell(row=r_idx, column=col_idx_s_rollover).value)
            expenses_val_s = to_decimal(ws_summary.cell(row=r_idx, column=col_idx_s_expenses).value)
            remaining_val_s = rollover_val_s + fy_contributions - fy_charges_offsets - expenses_val_s

            ws_summary.cell(row=r_idx, column=col_idx_s_contrib).value = float(fy_contributions)
            ws_summary.cell(row=r_idx, column=col_idx_s_charges).value = float(fy_charges_offsets)
            ws_summary.cell(row=r_idx, column=col_idx_s_remaining).value = float(remaining_val_s)

            for col_to_format in [col_idx_s_contrib, col_idx_s_charges, col_idx_s_expenses, col_idx_s_remaining]:
                 ws_summary.cell(row=r_idx, column=col_to_format).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        current_app.logger.info("Finished recalculating 'Summary' sheet.")

        # --- 5. Save the modified workbook ---
        workbook.save(temp_filepath)
        current_app.logger.info(f"Successfully updated and saved temporary summary file: {temp_filepath}")
        return True

    except Exception as e:
        current_app.logger.error(f"Error updating temporary summary file {temp_filepath} after review: {e}", exc_info=True)
        flash(f"Error updating the downloadable summary file after review: {e}", "danger")
        return False

@sc_bp.route('/foundation', methods=['GET', 'POST'])
@login_required
@admin_required
def foundation_reconciliation():
    form = FoundationReconciliationForm() 
    
    if form.validate_on_submit(): 
        session.pop('last_summary_temp_filename', None)
        session.pop('last_summary_timestamp', None)
        session.pop('last_results', None) 

        drs_file = form.drs_report.data
        donor_file = form.donor_report.data
        summary_file = form.summary_file.data

        if not drs_file or not donor_file or not summary_file:
            flash('Missing file data. Please select all three report files.', 'warning')
            return redirect(request.url)

        if (allowed_file(drs_file.filename) and
                allowed_file(donor_file.filename) and
                allowed_file(summary_file.filename)):

            upload_folder = os.path.join(current_app.instance_path, 'uploads_temp') 
            os.makedirs(upload_folder, exist_ok=True)
            timestamp_files = datetime.now().strftime('%Y%m%d%H%M%S%f')

            drs_filename_secure = secure_filename(f"drs_{current_user.id}_{timestamp_files}.xlsx")
            donor_filename_secure = secure_filename(f"donor_{current_user.id}_{timestamp_files}.xlsx")
            summary_filename_secure = secure_filename(f"summary_{current_user.id}_{timestamp_files}.xlsx")

            drs_path = os.path.join(upload_folder, drs_filename_secure)
            donor_path = os.path.join(upload_folder, donor_filename_secure)
            summary_path = os.path.join(upload_folder, summary_filename_secure)

            files_saved = False
            serializable_results_for_session = {} 

            try:
                drs_file.save(drs_path)
                donor_file.save(donor_path)
                summary_file.save(summary_path)
                files_saved = True
                flash('Files successfully uploaded. Processing...', 'info')
                current_app.logger.info("Input files saved temporarily, starting processing via update_summary_file.")

                file_buffer, results_from_processing = update_summary_file(drs_path, donor_path, summary_path) 
                current_app.logger.info(f"Raw results from update_summary_file: {results_from_processing}")
                
                actual_new_transactions_for_db = results_from_processing.get('new_transactions_for_db', [])

                serializable_results_for_session = {
                    'processed': results_from_processing.get('processed', 0),
                    'needs_review': results_from_processing.get('needs_review', 0),
                    'duplicates_db': results_from_processing.get('duplicates_db', 0),
                    'duplicates_sheet': results_from_processing.get('duplicates_sheet', 0),
                    'skipped_date_range': results_from_processing.get('skipped_date_range', 0),
                    'errors': results_from_processing.get('errors', []), 
                    'warnings': results_from_processing.get('warnings', []), 
                    'new_transactions_for_db_count': len(actual_new_transactions_for_db) 
                }
                session['last_results'] = serializable_results_for_session

                if serializable_results_for_session.get('errors'): 
                    for error in serializable_results_for_session['errors']: 
                        flash(f"Processing Error: {error}", 'danger')
                    current_app.logger.error(f"Processing errors: {serializable_results_for_session['errors']}")
                else: 
                    temp_file_storage_path = get_temp_file_storage_path() 
                    if file_buffer and temp_file_storage_path: 
                        temp_filename = f"summary_output_{uuid.uuid4().hex}.xlsx"
                        temp_filepath = os.path.join(temp_file_storage_path, temp_filename)
                        try:
                            with open(temp_filepath, 'wb') as f_temp:
                                f_temp.write(file_buffer.getvalue())
                            session['last_summary_temp_filename'] = temp_filename 
                            session['last_summary_timestamp'] = datetime.now().strftime("%Y-%m-%d %I:%M:%S %p")
                            current_app.logger.info(f"Stored updated summary in temporary file: {temp_filepath}")
                            flash('File processing complete. You can download the updated summary.', 'success')
                        except IOError as e:
                            current_app.logger.error(f"Could not write temp summary file: {temp_filepath}. Error: {e}")
                            flash("Error saving processed file. Download may not be available.", "danger")
                            if 'errors' not in serializable_results_for_session: serializable_results_for_session['errors'] = []
                            serializable_results_for_session['errors'].append("Server error: Could not save processed file.")
                            session['last_results'] = serializable_results_for_session 
                    elif not file_buffer:
                        if not serializable_results_for_session.get('warnings'): serializable_results_for_session['warnings'] = []
                        serializable_results_for_session['warnings'].append("Processing completed, but no update file was generated.")
                        session['last_results'] = serializable_results_for_session 
                        current_app.logger.warning("Processing completed but no summary buffer generated.")
                    elif not temp_file_storage_path: 
                        flash("Server config error: Temp storage not available. Download disabled.", "danger")
                        current_app.logger.error("Temp file storage path could not be accessed/created.")
                    
                    if serializable_results_for_session.get('warnings'):
                        for warning in serializable_results_for_session['warnings']: flash(f"Processing Warning: {warning}", 'warning')

                    if actual_new_transactions_for_db:
                        current_app.logger.info(f"Attempting to save {len(actual_new_transactions_for_db)} transactions to DB...")
                        with SessionLocal() as db_session: 
                            db_session.add_all(actual_new_transactions_for_db)
                            try:
                                db_session.commit()
                                current_app.logger.info("Successfully saved new transactions to DB.")
                                flash(f"{len(actual_new_transactions_for_db)} new transaction(s) saved to history.", "info")
                            except Exception as commit_err:
                                db_session.rollback()
                                current_app.logger.error(f"DB commit error: {commit_err}", exc_info=True)
                                flash(f"Error saving transaction history: {commit_err}", "danger")
                                if 'errors' not in serializable_results_for_session: serializable_results_for_session['errors'] = []
                                serializable_results_for_session['errors'].append(f"DB Save Error: {str(commit_err)}")
                                session['last_results'] = serializable_results_for_session 
            except Exception as e:
                flash(f'An error occurred during file upload or processing: {e}', 'danger')
                current_app.logger.error(f"Error during foundation upload/process: {e}", exc_info=True)
            finally:
                try: 
                    if files_saved:
                        if os.path.exists(drs_path): os.remove(drs_path)
                        if os.path.exists(donor_path): os.remove(donor_path)
                        if os.path.exists(summary_path): os.remove(summary_path)
                        current_app.logger.info("Temporary uploaded (input) files cleaned up.")
                except OSError as e:
                    current_app.logger.error(f"Error removing temporary uploaded (input) file: {e}")
            return redirect(url_for('.foundation_reconciliation'))
        else: 
            flash('Invalid file type. Only .xlsx files are allowed.', 'warning') 
            return redirect(request.url) 
    
    last_results_for_template = session.pop('last_results', None) if request.method == 'GET' else session.get('last_results')

    needs_review_count_db = 0
    try:
        with SessionLocal() as db_session:
            needs_review_count_db = db_session.query(FoundationTransaction).filter_by(status=FoundationTransaction.STATUS_NEEDS_REVIEW).count()
    except Exception as e:
        current_app.logger.error(f"Error querying needs_review_count: {e}", exc_info=True)
        flash("Could not retrieve count of items needing review.", "danger")

    return render_template(
        'sports_clubs/foundation_reconciliation.html',
        form=form, 
        processing_results=last_results_for_template, 
        num_needs_review_db=needs_review_count_db 
    )

@sc_bp.route('/foundation/manual_entry', methods=['GET', 'POST'])
@login_required
@admin_required
def foundation_manual_entry():
    form = ManualTransactionForm()
    try:
        with SessionLocal() as db_s:
            active_clubs_db = db_s.query(Club).filter_by(is_active=True).order_by(Club.name).all()
            # Choices for the form: (value_for_form_submission, display_text_in_dropdown)
            form.club_name.choices = [(club.id, club.name) for club in active_clubs_db]
            current_app.logger.info(f"Populated manual entry club dropdown with {len(active_clubs_db)} clubs from DB.")
    except Exception as e:
        current_app.logger.error(f"Error populating club choices for manual entry: {e}", exc_info=True)
        flash("Error loading club list for dropdown.", "danger")
        form.club_name.choices = [] # Ensure it's empty on error

    if not form.club_name.choices and request.method == 'GET': # Check only on GET if still no choices
        flash("No active clubs found in the database to populate the assignment dropdown. Please add clubs via admin if needed.", "warning")

    if form.validate_on_submit():
        if not session.get('last_summary_temp_filename'):
            flash("No active summary file to add this entry to. Please process files on the main reconciliation page first.", "danger")
            return redirect(url_for('.foundation_manual_entry'))

        selected_club_id = form.club_name.data # This is now the Club.id
        
        full_club_name_for_db_and_display = None
        sheet_target_club_name = None
        

        try:
            with SessionLocal() as db_s:
                assigned_club_obj = db_s.query(Club).get(selected_club_id)
                if not assigned_club_obj:
                    flash("Selected club not found in database. Please refresh.", "danger")
                    return redirect(url_for('.foundation_manual_entry'))
                full_club_name_for_db_and_display = assigned_club_obj.name
        except Exception as e:
            current_app.logger.error(f"Error fetching selected club (ID: {selected_club_id}) from DB: {e}", exc_info=True)
            flash("Error fetching club details.", "danger")
            return redirect(url_for('.foundation_manual_entry'))

        # Derive sheet target name (strip " Club")
        sheet_target_club_name = full_club_name_for_db_and_display
        if full_club_name_for_db_and_display.lower().endswith(" club"):
            sheet_target_club_name = full_club_name_for_db_and_display[:-5].strip()
        
        current_app.logger.info(f"Manual entry for DB Club: '{full_club_name_for_db_and_display}', Target Excel sheet name: '{sheet_target_club_name}'")

        # Prepare data for Excel helper function
        manual_data_for_excel = {
            'transaction_type': form.transaction_type.data,
            'transaction_date': form.transaction_date.data,
            'journal_ref': form.journal_ref.data,
            'name_description': form.name_description.data,
            'amount': form.amount.data, # This is a Decimal
            'club_name_for_sheet_operations': sheet_target_club_name, # Use the (potentially) shortened name for sheet ops
            'club_name_for_display': full_club_name_for_db_and_display, # Use full name for display in Excel row
            'notes': form.notes.data # This was previously removed due to DB error, ensure your model has it or remove if not
        }
        try:
            with SessionLocal() as db_s:
                existing_transaction = db_s.query(FoundationTransaction).filter(
                    FoundationTransaction.journal_ref == form.journal_ref.data
                ).first()
                if existing_transaction:
                    flash(f"Error: Journal Reference '{form.journal_ref.data}' already exists in the database. Please use a unique reference.", "danger")
                    # It's important to re-populate choices for the form if redirecting back to the same page on error
                    active_clubs_db_check = db_s.query(Club).filter_by(is_active=True).order_by(Club.name).all()
                    form.club_name.choices = [(club.id, club.name) for club in active_clubs_db_check]
                    return render_template('sports_clubs/manual_entry.html', form=form) # Re-render with form and error
        except Exception as e_check:
            current_app.logger.error(f"Error checking for existing journal ref: {e_check}", exc_info=True)
            flash("Error checking for existing transaction. Please try again.", "danger")
            return redirect(url_for('.foundation_manual_entry'))
        
        # Prepare data for DB
        gross_amt_db = Decimal('0.00')
        fees_total_db = Decimal('0.00')
        if form.transaction_type.data == 'Contribution':
            gross_amt_db = form.amount.data
        elif form.transaction_type.data == 'Expense':
            fees_total_db = form.amount.data 
        elif form.transaction_type.data == 'Refund': 
            fees_total_db = -form.amount.data

        new_db_transaction = FoundationTransaction(
            transaction_date=form.transaction_date.data,
            journal_ref=manual_data_for_excel['journal_ref'],
            donor_description=form.name_description.data,
            original_designation=f"Manual Entry: {form.transaction_type.data}", 
            gross_amount=gross_amt_db,
            fees_total=fees_total_db,
            net_amount=gross_amt_db - fees_total_db,
            assigned_club_name=full_club_name_for_db_and_display, # Store full DB name
            club_id=selected_club_id, # Store the Club's ID
            status=FoundationTransaction.STATUS_RECONCILED, 
            # notes=form.notes.data, # Add back ONLY if FoundationTransaction model has 'notes' field
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow()
        )
        assigned_club_obj = None
        try:
            with SessionLocal() as db_s:
                assigned_club_obj = db_s.query(Club).filter_by(id=selected_club_id).first()
                db_s.add(new_db_transaction)
                db_s.commit()
            flash(f"Manual entry '{form.name_description.data}' saved to database for {full_club_name_for_db_and_display}.", "info")
            current_app.logger.info(f"Manual transaction added to DB by {current_user.id}: {new_db_transaction.id}")

            if _update_temp_summary_after_review(new_db_transaction, assigned_club_obj): # Pass the dict with specific names
                session['last_summary_timestamp'] = datetime.now().strftime("%Y-%m-%d %I:%M:%S %p")
                flash("Manual entry added to downloadable summary file.", "success")
            else:
                flash("Manual entry saved to database, but failed to update the summary Excel file.", "warning")
            return redirect(url_for('.foundation_reconciliation')) 
        except Exception as e_db_manual:
            flash(f"Error saving manual entry to database: {e_db_manual}", "danger")
            current_app.logger.error(f"Error saving manual DB entry: {e_db_manual}", exc_info=True)
    
    return render_template('sports_clubs/manual_entry.html', form=form)

@sc_bp.route('/foundation/download_last')
@login_required
@admin_required
def download_last_summary():
    temp_filename = session.get('last_summary_temp_filename') 
    temp_file_storage_path = get_temp_file_storage_path(create=False) 

    if temp_filename and temp_file_storage_path:
        temp_filepath = os.path.join(temp_file_storage_path, temp_filename)
        if os.path.exists(temp_filepath):
            try:
                today = date.today()
                fiscal_year = today.year if today.month < 7 else today.year + 1
                fiscal_year_short = str(fiscal_year)[-2:]
                timestamp_str = session.get('last_summary_timestamp','timestamp_unavailable').replace(':','-').replace(' ','_')
                download_filename = f"FY{fiscal_year_short}_Foundation_Summary_Updated_{timestamp_str}.xlsx" 
                current_app.logger.info(f"Serving download: {download_filename} from {temp_filepath}")
                
                return send_file(
                    temp_filepath,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=download_filename
                )
            except Exception as e:
                current_app.logger.error(f"Error sending file {temp_filepath}: {e}", exc_info=True)
                flash("Error occurred while preparing file for download.", "danger")
        else:
            flash("Processed file not found. Please process files again.", "warning")
            current_app.logger.warning(f"Download attempt failed: temp file {temp_filepath} not found.")
            session.pop('last_summary_temp_filename', None) 
            session.pop('last_summary_timestamp', None)
    else:
        flash("No summary file available for download. Please process files first.", "warning")
        if not temp_filename: current_app.logger.warning("Download attempt: no temp filename in session.")
        if not temp_file_storage_path: current_app.logger.warning("Download attempt: temp file storage path not available.")
    
    return redirect(url_for('.foundation_reconciliation'))

@sc_bp.route('/foundation/clear_all_review_data', methods=['POST']) 
@login_required
@admin_required
def clear_all_review_data():
    """Clears ALL FoundationTransaction records from DB and the last generated temp summary file."""
    try:
        with SessionLocal() as db_session:
            num_deleted = db_session.query(FoundationTransaction).delete() # Delete ALL records
            db_session.commit()
            flash(f"{num_deleted} transaction record(s) cleared from the database.", "success")
            current_app.logger.info(f"All {num_deleted} FoundationTransaction items deleted from DB by user {current_user.id}.")
    except Exception as e:
        flash(f"Error clearing FoundationTransaction items from database: {e}", "danger")
        current_app.logger.error(f"Error clearing FoundationTransaction items from DB: {e}", exc_info=True)

    temp_filename_to_delete = session.pop('last_summary_temp_filename', None)
    session.pop('last_summary_timestamp', None) 
    session.pop('last_results', None) 

    if temp_filename_to_delete:
        temp_file_storage_path = get_temp_file_storage_path(create=False)
        if temp_file_storage_path:
            temp_filepath = os.path.join(temp_file_storage_path, temp_filename_to_delete)
            if os.path.exists(temp_filepath):
                try:
                    os.remove(temp_filepath)
                    flash(f"Temporary summary file '{temp_filename_to_delete}' deleted.", "info")
                    current_app.logger.info(f"Deleted temporary summary file: {temp_filepath}")
                except OSError as e:
                    flash(f"Error deleting temporary summary file '{temp_filename_to_delete}': {e}", "warning")
                    current_app.logger.error(f"Error deleting temporary summary file {temp_filepath}: {e}", exc_info=True)
        else:
            flash("Temporary file storage path not available, could not clear temp file.", "warning")
    else:
        flash("No active temporary summary file session to clear.", "info")
        
    return redirect(url_for('.foundation_reconciliation'))


@sc_bp.route('/foundation/review', methods=['GET', 'POST'])
@login_required
@admin_required
def foundation_review():
    if request.method == 'POST':
        transaction_id_str = request.form.get('transaction_id')
        club_id_str = request.form.get('club_id') 

        if not transaction_id_str or not club_id_str: 
            flash('Missing transaction ID or Club ID.', 'danger')
            return redirect(url_for('.foundation_review'))

        try:
            transaction_id = int(transaction_id_str)
            club_id = int(club_id_str) 

            with SessionLocal() as db_session: 
                transaction_to_review = db_session.query(FoundationTransaction).filter_by(id=transaction_id, status=FoundationTransaction.STATUS_NEEDS_REVIEW).first()
                assigned_club_obj = db_session.query(Club).filter_by(id=club_id).first()

                if not transaction_to_review:
                    flash('Transaction not found or already reviewed.', 'warning')
                elif not assigned_club_obj:
                    flash('Selected club not found.', 'danger')
                else:
                    transaction_to_review.club_id = assigned_club_obj.id
                    transaction_to_review.assigned_club_name = assigned_club_obj.name
                    transaction_to_review.status = FoundationTransaction.STATUS_RECONCILED 
                    transaction_to_review.updated_at = datetime.utcnow()
                    
                    db_session.commit() 
                    flash(f'Transaction {transaction_to_review.journal_ref or transaction_to_review.id} assigned to {assigned_club_obj.name}.', 'success')
                    current_app.logger.info(f"Transaction ID {transaction_id} assigned to Club ID {club_id} by User ID {current_user.id}.")
                    
                    if _update_temp_summary_after_review(transaction_to_review, assigned_club_obj):
                        session['last_summary_timestamp'] = datetime.now().strftime("%Y-%m-%d %I:%M:%S %p") 
                        flash('Downloadable summary file has been updated to reflect this review.', 'info')
                    else:
                        flash('Could not automatically update the downloadable summary file. It may not reflect this latest review.', 'warning')
        except ValueError: 
             flash('Invalid transaction or club ID format.', 'danger')
        except Exception as e:
            flash(f'An error occurred while assigning the club: {e}', 'danger')
            current_app.logger.error(f"Error assigning club for TX_ID {transaction_id_str}: {e}", exc_info=True) # Use str id for logging if int conversion failed
        
        return redirect(url_for('.foundation_review')) 

    transactions_to_review = []
    active_clubs = []
    try:
        with SessionLocal() as db_session: 
            transactions_to_review = db_session.query(FoundationTransaction).filter(
                FoundationTransaction.status == FoundationTransaction.STATUS_NEEDS_REVIEW
            ).order_by(FoundationTransaction.transaction_date.asc()).all() 
            active_clubs = db_session.query(Club).filter_by(is_active=True).order_by(Club.name).all()
    except Exception as e:
        current_app.logger.error(f"Error querying data for review page: {e}", exc_info=True)
        flash("Could not retrieve data for the review page.", "danger")

    return render_template(
        'sports_clubs/foundation_review.html',
        transactions=transactions_to_review,
        clubs=active_clubs
    )
