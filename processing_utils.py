# processing_utils.py
import pandas as pd
from decimal import Decimal, InvalidOperation
from datetime import datetime, date
import re
import sys
import traceback
import openpyxl # For reading/writing xlsx
from openpyxl.utils import get_column_letter # Already imported
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle, numbers
from io import BytesIO
from collections import defaultdict # For grouping transactions

from db import SessionLocal
# FoundationTransaction model is used for saving new entries to DB
from models import FoundationTransaction
from sqlalchemy import func

# --- Constants ---
DONOR_SHEET_NAME = 'College or Unit Acknowledgement'
DONOR_SKIP_ROWS = 7 # Header is row 8 in example
DRS_COL_DATE = 'Post Date'
DRS_COL_DESC_JRNL = 'Invoice Description / Journal Ref'
DRS_COL_DEBIT = 'Debit'
DRS_COL_CREDIT = 'Credit'
DRS_COL_TRANS_NUM = 'Transaction#'
# --- Donor Report Column Indices (0-based) ---
DONOR_COL_IDX_JRNL_REF = 1 # Column B = Journal Ref (Constituent ID)
DONOR_COL_IDX_DATE = 5     # Column F = Date
DONOR_COL_IDX_DESIGNATION = 8 # Column I = Designation Text

TARGET_SECTIONS = {
    "contributions - cash", "contributions - non cash",
    "services - bank/credit card fees", "services - cc platform processing fees",
    "transfer out - administrative gift fee"
}
SECTION_TO_TYPE = {
    "contributions - cash": "Contribution", "contributions - non cash": "Contribution",
    "services - bank/credit card fees": "Fee", "services - cc platform processing fees": "Fee",
    "transfer out - administrative gift fee": "Fee"
}
SUMMARY_SHEET_NAME = 'Summary'
SUMMARY_INDIVIDUAL_SHEET_NAME = 'Summary Individual' # New sheet
NEEDS_REVIEW_SHEET_NAME = 'Needs Review'

# --- Column Headers from User's Summary Example ---
SUMMARY_CLUB_COL = 'Sports Clubs'
SUMMARY_ROLLOVER_COL = 'Rollover'
SUMMARY_CONTRIB_COL = 'Sum of Contribution'
SUMMARY_CHARGES_COL = 'Sum of Chgs/offset' # Fees/Expenses from DRS
SUMMARY_EXPENSES_COL = 'Sum of Expenses' # Separate Expenses column (source TBD)
SUMMARY_REMAINING_COL = 'Sum of Remaining'

CLUB_SHEET_JRNL_COL = 'Journal Ref' # Column header for Journal Ref in individual club sheets
CLUB_SHEET_HEADERS = ['Date', 'Type', 'Journal Ref', 'Donor/Description', 'Contribution Amount', 'Charges/Offset', 'Net Amount', 'Donation Use']
NEEDS_REVIEW_HEADERS = ['Date', 'Type', 'Journal Ref', 'Donor/Description', 'Contribution Amount', 'Charges/Offset', 'Net Amount', 'Original Designation'] # Specific headers for this sheet
CLUB_SHEET_DATE_COL_IDX = 1
CLUB_SHEET_CONTRIB_COL_IDX = 5
CLUB_SHEET_EXPENSE_COL_IDX = 6 # Charges/Offset

# --- Helper to safely convert to Decimal ---
def to_decimal(value, make_positive=False, default=Decimal('0.00')):
    """Safely converts a value to Decimal, handling various inputs."""
    if pd.isna(value) or value is None or str(value).strip() == '': return default
    try:
        cleaned_value = str(value).replace(',', '').replace('(', '-').replace(')', '')
        if not re.match(r'^-?\d+(\.\d+)?$', cleaned_value): return default
        decimal_val = Decimal(cleaned_value)
        return abs(decimal_val) if make_positive else decimal_val
    except (InvalidOperation, ValueError, TypeError):
        return default

# --- Helper to Normalize Journal Refs ---
def normalize_jrnl_ref(ref):
    """Normalizes Journal Ref to a string, removing leading zeros if numeric."""
    if ref is None: return None
    ref_str = str(ref).strip()
    try:
        return str(int(ref_str)) 
    except ValueError:
        return ref_str.upper() 

# --- Helper to Normalize Text for Matching ---
def normalize_text_for_matching(text):
    """Lowercase, remove apostrophes/quotes, replace special chars with space."""
    if not text: return ""
    text = str(text).lower()
    text = text.replace("'", "").replace("’", "").replace('"', "").replace('`', '')
    text = re.sub(r'[^\w\s]+', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

# --- Club Name Matching (using list from summary and normalized text) ---
def find_matching_club_sheet(designation_text, club_names_from_summary):
    """Finds the best matching club name from the summary list based on designation."""
    if not designation_text or pd.isna(designation_text): return None
    normalized_designation = normalize_text_for_matching(designation_text)
    if not normalized_designation: return None

    potential_matches = []
    normalized_to_original_club = {normalize_text_for_matching(c): c for c in club_names_from_summary}

    for normalized_club_name, original_club_name in normalized_to_original_club.items():
        pattern = r'\b' + re.escape(normalized_club_name) + r'\b'
        if re.search(pattern, normalized_designation):
            potential_matches.append({'name': original_club_name, 'match_len': len(normalized_club_name)})

    if not potential_matches: return None
    elif len(potential_matches) == 1:
        return potential_matches[0]['name']
    else:
        potential_matches.sort(key=lambda x: x['match_len'], reverse=True)
        if potential_matches[0]['match_len'] > potential_matches[1]['match_len']:
            return potential_matches[0]['name']
        else: 
            exact_match_original_names = [m['name'] for m in potential_matches if normalize_text_for_matching(m['name']) == normalized_designation]
            if len(exact_match_original_names) == 1:
                return exact_match_original_names[0]
            # If still ambiguous, log it (optional) and return None
            # ambiguous_names = [m['name'] for m in potential_matches if m['match_len'] == potential_matches[0]['match_len']]
            # print(f"[Debug Match] Ambiguous normalized match for '{designation_text}'. Potential: {ambiguous_names}")
            return None


# --- Main Processing Function ---
def update_summary_file(drs_file_path, donor_file_path, summary_file_path):
    """Processes DRS, Donor, and Summary Excel files to update the foundation summary."""
    print("\n" + "="*20 + " Starting Foundation Summary Update " + "="*20)
    results = {'processed': 0, 'needs_review': 0, 'duplicates_db': 0, 'duplicates_sheet': 0, 'skipped_date_range': 0, 'errors': []}
    new_transactions_for_db = []
    transactions_by_ref = defaultdict(lambda: {'date': None, 'lines': [], 'designation': None, 'raw_jrnl_ref': None})
    all_processed_transactions_for_summary_individual = [] 
    output_buffer = None
    workbook = None # Initialize workbook variable
    min_donor_date = None
    max_donor_date = None

    try:
        # --- 1. Read Donor Report ---
        print(f"[Step 1] Reading Donor Report: {donor_file_path}")
        jrnl_to_designation = {} 
        donor_dates = []
        try:
            # Read relevant columns from Donor Report
            donor_df = pd.read_excel(
                donor_file_path,
                sheet_name=DONOR_SHEET_NAME,
                skiprows=DONOR_SKIP_ROWS,
                header=None, # Read without header initially to handle potential variations
                usecols=[DONOR_COL_IDX_JRNL_REF, DONOR_COL_IDX_DATE, DONOR_COL_IDX_DESIGNATION],
                names=['RawJrnlRef', 'Date', 'DesignationText'] # Assign temporary names
            )
            # Drop rows where any of these key fields are missing
            donor_df = donor_df.dropna(subset=['RawJrnlRef', 'DesignationText', 'Date']) 
            donor_df['DesignationText'] = donor_df['DesignationText'].astype(str).str.strip()

            # Build map using normalized Jrnl Ref and collect dates
            for _, row in donor_df.iterrows():
                 raw_ref = row['RawJrnlRef']
                 designation_text = row['DesignationText']
                 normalized_ref = normalize_jrnl_ref(raw_ref) # Normalize the journal reference
                 if normalized_ref:
                     # Store designation only if not already present for this normalized ref (first one wins)
                     if normalized_ref not in jrnl_to_designation:
                        jrnl_to_designation[normalized_ref] = designation_text
                 
                 # Collect and validate dates
                 try:
                     donor_date = pd.to_datetime(row['Date'], errors='coerce').date() # Convert to date object
                     if pd.notna(donor_date): # Check if conversion was successful
                         donor_dates.append(donor_date)
                 except Exception as date_err: # Catch any error during date parsing for a specific row
                     print(f"  [Warning] Error parsing date '{row['Date']}' in Donor Report: {date_err}") 

            if donor_dates:
                min_donor_date = min(donor_dates)
                max_donor_date = max(donor_dates)
                print(f"[Step 1] Donor Report Date Range: {min_donor_date} to {max_donor_date}")
            else:
                 print("[Warning] No valid dates found in Donor Report. Date range filtering will be skipped.")
            print(f"[Step 1] Created Journal Ref -> Designation map with {len(jrnl_to_designation)} entries.")
        except Exception as e: # Catch broader errors in reading the Donor Report
            print(f"[Warning] Error reading or parsing Donor Report: {e}. Lookup/filtering might be incomplete.")
            jrnl_to_designation = {} # Ensure it's an empty dict on error

        # --- 2. Read Existing Summary File ---
        print(f"[Step 2] Reading existing Summary file: {summary_file_path}")
        clubs_from_summary = []
        summary_header_map = {}
        summary_header_row_idx = -1 # Initialize to -1 (not found)
        try:
            workbook = openpyxl.load_workbook(summary_file_path) # Load the workbook
            if SUMMARY_SHEET_NAME not in workbook.sheetnames: # Check if 'Summary' sheet exists
                 raise ValueError(f"'{SUMMARY_SHEET_NAME}' sheet not found in the uploaded summary file.")
            ws_summary = workbook[SUMMARY_SHEET_NAME] # Get the 'Summary' worksheet

            # Define required headers for the 'Summary' sheet
            required_summary_headers = [SUMMARY_CLUB_COL, SUMMARY_ROLLOVER_COL, SUMMARY_CONTRIB_COL, SUMMARY_CHARGES_COL, SUMMARY_EXPENSES_COL, SUMMARY_REMAINING_COL]
            # Find the header row in the 'Summary' sheet
            for row_idx in range(1, min(20, ws_summary.max_row + 1)): # Search in first 20 rows
                row_values = [str(cell.value).strip() if cell.value is not None else "" for cell in ws_summary[row_idx]]
                if all(req_h in row_values for req_h in required_summary_headers): # Check if all required headers are in this row
                    summary_header_row_idx = row_idx
                    for col_idx, header_val in enumerate(row_values): # Map header names to column indices
                        if header_val: summary_header_map[header_val] = col_idx + 1
                    break
            if summary_header_row_idx == -1: raise ValueError(f"Header row containing required columns not found in '{SUMMARY_SHEET_NAME}' sheet.")
            missing_headers = [h for h in required_summary_headers if h not in summary_header_map]
            if missing_headers: raise ValueError(f"Missing required columns in '{SUMMARY_SHEET_NAME}' sheet: {missing_headers}")

            # Get club names from the 'Summary' sheet
            col_idx_club_summary = summary_header_map[SUMMARY_CLUB_COL]
            for row_idx in range(summary_header_row_idx + 1, ws_summary.max_row + 1): # Start from row after header
                 cell_value = ws_summary.cell(row=row_idx, column=col_idx_club_summary).value
                 if cell_value: # If cell is not empty
                     club_name = str(cell_value).strip()
                     if club_name and club_name.lower() not in ['grand total']: # Exclude 'grand total'
                         clubs_from_summary.append(club_name)
            if not clubs_from_summary: raise ValueError(f"No club names found below the header in column '{SUMMARY_CLUB_COL}' of the '{SUMMARY_SHEET_NAME}' sheet.")
            print(f"[Step 2] Found {len(clubs_from_summary)} clubs in Summary sheet.")
            print(f"[Step 2] Loaded existing summary workbook. Sheets: {workbook.sheetnames}")

            # Ensure sheets exist for all clubs and standard sheets
            print("  Ensuring sheets exist for all clubs listed in Summary...")
            sheets_to_ensure = clubs_from_summary + [NEEDS_REVIEW_SHEET_NAME, SUMMARY_INDIVIDUAL_SHEET_NAME]
            for sheet_name_to_create in sheets_to_ensure:
                # Sanitize sheet name (max 31 chars, no invalid chars)
                safe_sheet_name = re.sub(r'[\\/*?:\[\]]', '_', sheet_name_to_create)[:31]
                if safe_sheet_name not in workbook.sheetnames:
                    ws_new = workbook.create_sheet(safe_sheet_name)
                    headers_to_use = NEEDS_REVIEW_HEADERS if safe_sheet_name == NEEDS_REVIEW_SHEET_NAME else CLUB_SHEET_HEADERS
                    ws_new.append(headers_to_use)
                    for cell in ws_new[1]: cell.font = Font(bold=True) # Bold header
                    ws_new.freeze_panes = 'A2' # Freeze header row
                    print(f"    Created missing sheet: '{safe_sheet_name}'")
        except Exception as e: # Catch errors during summary file reading/preparation
            error_msg = f"Error reading/preparing existing Summary file: {e}"; results['errors'].append(error_msg); print(f"[Error] {error_msg}"); traceback.print_exc(); return None, results

        # --- 3. Get Existing Journal Refs from DB ---
        print("[Step 3] Fetching existing Journal Refs from database...")
        with SessionLocal() as db: # Use SQLAlchemy session
            existing_refs_db_normalized = {normalize_jrnl_ref(ref[0]) for ref in db.query(FoundationTransaction.journal_ref).all()}
        print(f"[Step 3] Found {len(existing_refs_db_normalized)} unique normalized transaction references in DB.")

        # --- 4. Read DRS Report ---
        drs_sheet_to_read = None # Initialize sheet name
        try:
            xls = pd.ExcelFile(drs_file_path) # Read Excel file object
            # Try to find a relevant sheet name
            potential_sheets = [name for name in xls.sheet_names if 'activity' in name.lower() or 'ledger' in name.lower()]
            if potential_sheets: drs_sheet_to_read = potential_sheets[0] # Take the first potential match
            elif '4100-774390' in xls.sheet_names: drs_sheet_to_read = '4100-774390' # Specific default
            else: drs_sheet_to_read = xls.sheet_names[0] # Fallback to the first sheet
            print(f"[Step 4] Reading DRS Report: {drs_file_path}, Selected Sheet: '{drs_sheet_to_read}'")
            drs_df_raw = pd.read_excel(xls, sheet_name=drs_sheet_to_read, header=None) # Read without header
            print(f"[Step 4] Read {len(drs_df_raw)} raw rows from DRS Report sheet.")
        except Exception as e: # Catch errors during DRS file reading
            error_msg = f"Error reading DRS Report sheet '{drs_sheet_to_read or 'unknown'}': {e}"; results['errors'].append(error_msg); print(f"[Error] {error_msg}"); return None, results

        # --- 5. Parse DRS Rows ---
        print("[Step 5] Parsing DRS rows into intermediate structure...")
        current_section_name = None; in_target_section = False; actual_headers = {}; header_row_index = -1
        # Define header candidates to look for in DRS report
        header_candidates = [DRS_COL_DATE, DRS_COL_DEBIT, DRS_COL_CREDIT, DRS_COL_DESC_JRNL]
        # Find the header row in DRS report
        for idx, row in drs_df_raw.iterrows():
            row_values_stripped = {str(v).strip() for v in row.tolist() if pd.notna(v)} # Get unique, stripped values from row
            if all(header in row_values_stripped for header in header_candidates): # Check if all candidates are present
                header_row_index = idx
                for col_idx, header_val in enumerate(row.tolist()): # Map header names to column indices
                    if pd.notna(header_val): actual_headers[str(header_val).strip()] = col_idx
                print(f"  Found DRS header row at index {header_row_index}.")
                break
        if header_row_index == -1: results['errors'].append(f"Could not find header row in DRS Report."); print(f"[Error] Could not find DRS header row."); return None, results
        # Check for required column keys in the identified header
        required_keys = [DRS_COL_DATE, DRS_COL_DESC_JRNL, DRS_COL_DEBIT, DRS_COL_CREDIT]
        if not all(key in actual_headers for key in required_keys):
            missing = [key for key in required_keys if key not in actual_headers]; results['errors'].append(f"DRS Header row missing required columns: {missing}."); print(f"[Error] DRS Header row missing required columns: {missing}"); return None, results
        # Get column indices from the header map
        col_idx_date = actual_headers[DRS_COL_DATE]; col_idx_desc_jrnl = actual_headers[DRS_COL_DESC_JRNL]; col_idx_debit = actual_headers[DRS_COL_DEBIT]; col_idx_credit = actual_headers[DRS_COL_CREDIT]; col_idx_trans_num = actual_headers.get(DRS_COL_TRANS_NUM)

        # Iterate through DRS rows to populate transactions_by_ref
        for index, row_data in drs_df_raw.iterrows():
            row_list = row_data.tolist() # Convert row to list
            first_cell_value = str(row_list[0]).strip() if pd.notna(row_list[0]) else None # Get first cell value
            # Identify various types of non-data rows
            is_blank_row = all(pd.isna(v) for v in row_list)
            is_total_row = first_cell_value and "total" in first_cell_value.lower() and "grand total" not in first_cell_value.lower()
            is_grand_total_row = first_cell_value and "grand total" in first_cell_value.lower()
            is_section_header_row = False; is_drs_header_row = (index == header_row_index)

            # Section Detection Logic (based on "Account:")
            if first_cell_value and first_cell_value.startswith("Account:"):
                is_section_header_row = True
                match = re.match(r"Account:\s*\S+\s*\((.*?)\)?$", first_cell_value) # Regex to extract section name
                potential_section_name = match.group(1).strip().lower() if match else first_cell_value.split(":", 1)[-1].strip().lower()
                if potential_section_name in TARGET_SECTIONS: # Check if it's a target section
                    if current_section_name != potential_section_name: print(f"\n[Section Change] Entered target section: '{potential_section_name}' at row index {index}")
                    current_section_name = potential_section_name; in_target_section = True
                else: # Not a target section
                    if in_target_section: print(f"  [Section End] Exiting section '{current_section_name}' at index {index}.")
                    in_target_section = False; current_section_name = None
                continue # Skip the section header row itself
            
            # Skip blank, total, grand total, or DRS header rows
            if is_blank_row or is_total_row or is_grand_total_row or is_drs_header_row:
                if in_target_section and (is_blank_row or is_total_row or is_grand_total_row): # If exiting a section due to these rows
                    print(f"  [Section End] Exiting section '{current_section_name}' at index {index}.")
                    in_target_section = False; current_section_name = None;
                if is_grand_total_row: break # Stop processing at grand total
                continue
            
            # Skip if not currently in a target section
            if not in_target_section or not current_section_name: continue

            # Process Data Row
            try:
                trans_date_raw = row_list[col_idx_date]
                combined_desc_ref = str(row_list[col_idx_desc_jrnl]) if pd.notna(row_list[col_idx_desc_jrnl]) else ""
                debit_val = to_decimal(row_list[col_idx_debit])
                credit_val = to_decimal(row_list[col_idx_credit], make_positive=True) # Contributions are positive
                trans_num_raw = str(row_list[col_idx_trans_num]).strip() if col_idx_trans_num is not None and pd.notna(row_list[col_idx_trans_num]) else None
                trans_date = pd.to_datetime(trans_date_raw, errors='coerce').date() # Convert to date
                if not trans_date: continue # Skip if date is invalid

                # Apply Date Range Filter
                if min_donor_date and max_donor_date: # Only if donor dates are valid
                    try:
                        if not (min_donor_date <= trans_date <= max_donor_date):
                            # Check if this ref was already skipped to avoid double counting skipped_date_range
                            temp_raw_ref = None
                            match_ref_temp = re.search(r'[ -]?([a-zA-Z0-9]{3,})$', combined_desc_ref) # Try to get ref
                            if match_ref_temp: temp_raw_ref = match_ref_temp.group(1).strip()
                            elif trans_num_raw and re.search(r'\d', trans_num_raw): temp_raw_ref = trans_num_raw
                            if temp_raw_ref and normalize_jrnl_ref(temp_raw_ref) not in transactions_by_ref: 
                                results['skipped_date_range'] += 1
                            continue # Skip DRS rows outside the Donor report's date range
                    except TypeError as e:
                        error_message = f"Error processing DRS row {idx + header_row_index + 1} (Date: {row.get(date_col_header)}): {e}. Check date format."
                        logger.error(error_message, exc_info=False) # Log less verbosely here if needed
                        results['errors'].append(error_message) # IMPORTANT
                        results['skipped_other_error'] = results.get('skipped_other_error', 0) + 1
                        continue # Skip this transaction
                
                # Extract Journal Ref from DRS row
                raw_jrnl_ref = None; match_ref = re.search(r'[ -]?([a-zA-Z0-9]{3,})$', combined_desc_ref) # Regex for ref at end
                if match_ref: raw_jrnl_ref = match_ref.group(1).strip()
                elif trans_num_raw and re.search(r'\d', trans_num_raw): raw_jrnl_ref = trans_num_raw # Fallback to trans_num
                else: # Further fallback for complex descriptions
                    parts = combined_desc_ref.split('/'); jrnl_ref_part = parts[-1].strip().replace('\xa0', '').strip() if len(parts) > 1 else ""
                    if re.search(r'\d', jrnl_ref_part) and len(jrnl_ref_part) > 4: raw_jrnl_ref = jrnl_ref_part
                if not raw_jrnl_ref: continue # Skip if no ref found
                normalized_jrnl_ref = normalize_jrnl_ref(raw_jrnl_ref) # Normalize it
                if not normalized_jrnl_ref: continue

                # Determine line type and amount
                line_type = SECTION_TO_TYPE.get(current_section_name); line_amount = Decimal('0.00'); line_desc = combined_desc_ref
                if line_type == "Contribution": line_amount = credit_val
                elif line_type == "Fee": line_amount = debit_val # Fees are positive debits
                else: continue # Not a relevant line type
                if line_amount == 0: continue # Skip zero-value lines

                # Store or update the transaction details using NORMALIZED ref as key
                if not transactions_by_ref[normalized_jrnl_ref]['date']: # If first time seeing this ref
                     transactions_by_ref[normalized_jrnl_ref]['date'] = trans_date
                     transactions_by_ref[normalized_jrnl_ref]['raw_jrnl_ref'] = raw_jrnl_ref # Store original format too
                transactions_by_ref[normalized_jrnl_ref]['lines'].append({'type': line_type, 'amount': line_amount, 'desc': line_desc})
                # Store original designation from donor report map (using normalized ref)
                if normalized_jrnl_ref not in transactions_by_ref or transactions_by_ref[normalized_jrnl_ref].get('designation') is None:
                     transactions_by_ref[normalized_jrnl_ref]['designation'] = jrnl_to_designation.get(normalized_jrnl_ref) 
            except Exception as row_err: # Catch errors during individual row processing
                 print(f"  [Error] Processing DRS row {index}: {row_err}"); traceback.print_exc()
        print(f"[Step 5] Parsed {len(transactions_by_ref)} potential transactions by Journal Ref from DRS (filtered by date).")
        if results['skipped_date_range'] > 0: print(f"  Skipped {results['skipped_date_range']} DRS transactions outside Donor date range.")

        # --- 6. Process Aggregated Transactions & Update Workbook ---
        print("[Step 6] Aggregating transactions and updating workbook sheets...")
        ws_needs_review = workbook[NEEDS_REVIEW_SHEET_NAME]
        ws_summary_individual = workbook[SUMMARY_INDIVIDUAL_SHEET_NAME]

        for normalized_jrnl_ref, data in transactions_by_ref.items():
            # Aggregate amounts for this journal reference
            total_contribution = sum(line['amount'] for line in data['lines'] if line['type'] == 'Contribution')
            total_fees = sum(line['amount'] for line in data['lines'] if line['type'] == 'Fee')
            net_amount = total_contribution - total_fees; trans_date = data['date']; original_designation = data['designation']; raw_jrnl_ref = data['raw_jrnl_ref']
            
            # Determine primary description from transaction lines
            primary_desc = "[Desc N/A]"; contrib_lines = [l for l in data['lines'] if l['type'] == 'Contribution']; fee_lines = [l for l in data['lines'] if l['type'] == 'Fee']
            if contrib_lines: # Prefer contribution description
                raw_desc = contrib_lines[0]['desc']; match_desc = re.match(r"(?:Cash Contributions?|GIFT RECEIVED|DONATION)\s*(?:-|from)?\s*(.*)", raw_desc, re.IGNORECASE); donor_part = raw_desc
                if raw_jrnl_ref and raw_jrnl_ref in donor_part: donor_part = donor_part.replace(f"/{raw_jrnl_ref}", "").replace(f"-{raw_jrnl_ref}", "").replace(raw_jrnl_ref, "").strip(" /").strip() # Clean out ref
                primary_desc = match_desc.group(1).strip() if match_desc and match_desc.group(1).strip() else re.sub(r"^(Cash Contribution|GIFT RECEIVED|DONATION)\s*[-]?\s*", "", donor_part, flags=re.IGNORECASE).strip()
                if not primary_desc: primary_desc = "[Donor Name N/A]" # Fallback if parsing fails
            elif fee_lines: # Use fee description if no contribution
                raw_desc = fee_lines[0]['desc']
                if "ADMINISTRATIVE GIFT FEE" in raw_desc.upper(): primary_desc = "OSU Foundation Fee"
                elif "CC PLATFORM PROCESSING FEES" in raw_desc.upper(): primary_desc = "Credit Card Platform Fee"
                elif "BANK/CREDIT CARD FEES" in raw_desc.upper(): primary_desc = "Bank/Credit Card Fee"
                else: primary_desc = f"[Fee] {raw_desc}"
            else: primary_desc = data['lines'][0]['desc'] if data['lines'] else "[Unknown Desc]" # Fallback
            
            # Match Club using Summary List and the fetched Designation
            matched_club_name = find_matching_club_sheet(original_designation, clubs_from_summary)
            
            # Prepare data for Excel row and DB model
            excel_data_dict = {'Date': trans_date, 'Type': "Contribution" if total_contribution > 0 else "Fee/Expense", 'Journal Ref': raw_jrnl_ref, 'Donor/Description': primary_desc, 'Contribution Amount': total_contribution, 'Charges/Offset': total_fees, 'Net Amount': net_amount, 'Donation Use': matched_club_name or ""}
            db_model_data = {'transaction_date': trans_date, 'journal_ref': raw_jrnl_ref, 'donor_description': primary_desc, 'original_designation': original_designation, 'gross_amount': total_contribution, 'fees_total': total_fees, 'net_amount': net_amount, 'assigned_club_name': matched_club_name}
            
            # Determine target sheet and DB status
            target_sheet_name = None; is_needs_review = False
            if matched_club_name:
                safe_sheet_name = re.sub(r'[\\/*?:\[\]]', '_', matched_club_name)[:31] # Sanitize club name for sheet
                if safe_sheet_name in workbook.sheetnames: target_sheet_name = safe_sheet_name; db_model_data['status'] = FoundationTransaction.STATUS_RECONCILED
                else: print(f"  [Error] Matched club '{matched_club_name}' but sheet '{safe_sheet_name}' not found. Adding to Needs Review."); target_sheet_name = NEEDS_REVIEW_SHEET_NAME; is_needs_review = True; db_model_data['status'] = FoundationTransaction.STATUS_NEEDS_REVIEW
            else: target_sheet_name = NEEDS_REVIEW_SHEET_NAME; is_needs_review = True; db_model_data['status'] = FoundationTransaction.STATUS_NEEDS_REVIEW
            
            # Check if JrnlRef already exists in the target Excel sheet
            ws_target = workbook[target_sheet_name]; jrnl_col_idx_target = -1
            for col_idx, cell in enumerate(ws_target[1]): # Find 'Journal Ref' column in header
                if cell.value == CLUB_SHEET_JRNL_COL: jrnl_col_idx_target = col_idx + 1; break
            jrnl_exists_in_sheet = False
            if jrnl_col_idx_target != -1: # If 'Journal Ref' column found
                for row_idx_check in range(2, ws_target.max_row + 1): # Check data rows
                    if normalize_jrnl_ref(ws_target.cell(row=row_idx_check, column=jrnl_col_idx_target).value) == normalized_jrnl_ref: jrnl_exists_in_sheet = True; break
            
            if not jrnl_exists_in_sheet: # If not a duplicate in the sheet
                # Prepare row values based on target sheet headers
                new_row_values = [excel_data_dict.get(h) for h in (NEEDS_REVIEW_HEADERS if target_sheet_name == NEEDS_REVIEW_SHEET_NAME else CLUB_SHEET_HEADERS)]
                if target_sheet_name == NEEDS_REVIEW_SHEET_NAME: new_row_values[-1] = original_designation # Ensure original designation is in the last col for Needs Review
                else: all_processed_transactions_for_summary_individual.append(excel_data_dict) # Add to list for 'Summary Individual'
                
                ws_target.append(new_row_values); new_row_num = ws_target.max_row # Append and get new row number
                # Apply formatting to the new row
                header_map_target = {str(cell.value).strip(): c_idx+1 for c_idx, cell in enumerate(ws_target[1]) if cell.value}
                date_col_idx_fmt = header_map_target.get('Date'); contrib_col_idx_fmt = header_map_target.get('Contribution Amount'); charges_col_idx_fmt = header_map_target.get('Charges/Offset'); net_col_idx_fmt = header_map_target.get('Net Amount')
                if date_col_idx_fmt: ws_target.cell(row=new_row_num, column=date_col_idx_fmt).number_format = numbers.FORMAT_DATE_YYYYMMDD2
                if contrib_col_idx_fmt: ws_target.cell(row=new_row_num, column=contrib_col_idx_fmt).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                if charges_col_idx_fmt: ws_target.cell(row=new_row_num, column=charges_col_idx_fmt).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                if net_col_idx_fmt: ws_target.cell(row=new_row_num, column=net_col_idx_fmt).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                
                # Add to DB list if it wasn't already in the DB
                if normalized_jrnl_ref not in existing_refs_db_normalized: new_transactions_for_db.append(FoundationTransaction(**db_model_data))
                if is_needs_review: results['needs_review'] += 1 # Increment 'needs_review' count
                results['processed'] += 1 # Increment processed count
            else: results['duplicates_sheet'] += 1 # Increment duplicate sheet count

        # --- Populate Summary Individual Sheet ---
        print(f"  Populating '{SUMMARY_INDIVIDUAL_SHEET_NAME}' sheet...")
        ws_summary_individual = workbook[SUMMARY_INDIVIDUAL_SHEET_NAME]
        if ws_summary_individual.max_row > 1: ws_summary_individual.delete_rows(2, ws_summary_individual.max_row - 1) # Clear existing data
        all_processed_transactions_for_summary_individual.sort(key=lambda x: (x['Date'] if x['Date'] else date.min, x['Donation Use'] if x['Donation Use'] else "")) # Sort for consistent order
        for trans_dict in all_processed_transactions_for_summary_individual:
            new_row_values = [trans_dict.get(h) for h in CLUB_SHEET_HEADERS] # Get values based on headers
            ws_summary_individual.append(new_row_values); new_row_num = ws_summary_individual.max_row
            # Apply formatting
            header_map_indiv = {str(cell.value).strip(): c_idx+1 for c_idx, cell in enumerate(ws_summary_individual[1]) if cell.value}
            date_col_idx_fmt = header_map_indiv.get('Date'); contrib_col_idx_fmt = header_map_indiv.get('Contribution Amount'); charges_col_idx_fmt = header_map_indiv.get('Charges/Offset'); net_col_idx_fmt = header_map_indiv.get('Net Amount')
            if date_col_idx_fmt: ws_summary_individual.cell(row=new_row_num, column=date_col_idx_fmt).number_format = numbers.FORMAT_DATE_YYYYMMDD2
            if contrib_col_idx_fmt: ws_summary_individual.cell(row=new_row_num, column=contrib_col_idx_fmt).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            if charges_col_idx_fmt: ws_summary_individual.cell(row=new_row_num, column=charges_col_idx_fmt).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            if net_col_idx_fmt: ws_summary_individual.cell(row=new_row_num, column=net_col_idx_fmt).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

        # --- 7. Recalculate Summary Sheet ---
        print("[Step 7] Recalculating Summary sheet...")
        ws_summary = workbook[SUMMARY_SHEET_NAME] 
        # Re-verify summary headers to get column indices
        summary_headers = {}; summary_header_row_idx = -1 
        required_summary_headers = [SUMMARY_CLUB_COL, SUMMARY_ROLLOVER_COL, SUMMARY_CONTRIB_COL, SUMMARY_CHARGES_COL, SUMMARY_EXPENSES_COL, SUMMARY_REMAINING_COL]
        for row_idx in range(1, ws_summary.max_row + 1): # Search all rows for header
            row_values = [str(cell.value).strip() if cell.value is not None else "" for cell in ws_summary[row_idx]]
            if all(req_h in row_values for req_h in required_summary_headers):
                summary_header_row_idx = row_idx
                for col_idx, header_val in enumerate(row_values):
                    if header_val: summary_headers[header_val] = col_idx + 1
                break
        if summary_header_row_idx == -1 or not all(h in summary_headers for h in required_summary_headers):
             results['errors'].append(f"Could not re-verify required headers in '{SUMMARY_SHEET_NAME}' for recalculation."); return None, results
        # Get column indices from verified headers
        col_idx_club = summary_headers[SUMMARY_CLUB_COL]; col_idx_rollover = summary_headers[SUMMARY_ROLLOVER_COL]; col_idx_contrib = summary_headers[SUMMARY_CONTRIB_COL]; col_idx_charges = summary_headers[SUMMARY_CHARGES_COL]; col_idx_expenses = summary_headers[SUMMARY_EXPENSES_COL]; col_idx_remaining = summary_headers[SUMMARY_REMAINING_COL]
        today = date.today(); current_fiscal_year_start = date(today.year if today.month >= 7 else today.year - 1, 7, 1) # Fiscal year starts July 1st

        # Iterate through summary rows and update totals
        for row_idx in range(summary_header_row_idx + 1, ws_summary.max_row + 1): # Start after header
            club_name_cell = ws_summary.cell(row=row_idx, column=col_idx_club); club_name = club_name_cell.value
            if not club_name or str(club_name).strip().lower() in ['grand total']: continue # Skip empty or grand total rows
            club_name = str(club_name).strip(); safe_sheet_name = re.sub(r'[\\/*?:\[\]]', '_', club_name)[:31] # Sanitize name
            fy_contrib_total = Decimal('0.00'); fy_charges_total = Decimal('0.00') 
            if safe_sheet_name in workbook.sheetnames: # If club sheet exists
                ws_club = workbook[safe_sheet_name]; club_sheet_header_map = {}
                if ws_club.max_row > 1: # If sheet has data
                    club_sheet_header_row = ws_club[1] # Assume header is row 1
                    for col_idx, cell in enumerate(club_sheet_header_row): # Map club sheet headers
                        if cell.value: club_sheet_header_map[str(cell.value).strip()] = col_idx + 1
                    # Get column indices for Date, Contribution, Charges
                    date_col_idx = club_sheet_header_map.get('Date'); contrib_col_idx = club_sheet_header_map.get('Contribution Amount'); charges_col_idx = club_sheet_header_map.get('Charges/Offset') 
                    if date_col_idx and contrib_col_idx and charges_col_idx: # If all needed columns found
                        for club_row_idx in range(2, ws_club.max_row + 1): # Iterate data rows
                            trans_date_val = ws_club.cell(row=club_row_idx, column=date_col_idx).value; trans_date_obj = None
                            # Convert date robustly
                            if isinstance(trans_date_val, datetime): trans_date_obj = trans_date_val.date()
                            elif isinstance(trans_date_val, date): trans_date_obj = trans_date_val
                            elif isinstance(trans_date_val, (int, float)): # Handle Excel serial dates
                                try: trans_date_obj = pd.to_datetime(trans_date_val, unit='D', origin='1899-12-30').date()
                                except: pass
                            elif isinstance(trans_date_val, str): # Handle string dates
                                try: trans_date_obj = pd.to_datetime(trans_date_val).date()
                                except: pass
                            # Sum up contributions and charges within the current fiscal year
                            if trans_date_obj and trans_date_obj >= current_fiscal_year_start:
                                fy_contrib_total += to_decimal(ws_club.cell(row=club_row_idx, column=contrib_col_idx).value)
                                fy_charges_total += to_decimal(ws_club.cell(row=club_row_idx, column=charges_col_idx).value)
            # Get rollover and expenses from summary sheet
            rollover_val = to_decimal(ws_summary.cell(row=row_idx, column=col_idx_rollover).value); expenses_val = to_decimal(ws_summary.cell(row=row_idx, column=col_idx_expenses).value)
            # Calculate remaining amount
            current_remaining_val = rollover_val + fy_contrib_total - fy_charges_total - expenses_val
            # Update summary sheet cells with new totals
            ws_summary.cell(row=row_idx, column=col_idx_contrib).value = float(fy_contrib_total); ws_summary.cell(row=row_idx, column=col_idx_charges).value = float(fy_charges_total); ws_summary.cell(row=row_idx, column=col_idx_expenses).value = float(expenses_val); ws_summary.cell(row=row_idx, column=col_idx_remaining).value = float(current_remaining_val)
            # Apply currency formatting
            ws_summary.cell(row=row_idx, column=col_idx_contrib).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE; ws_summary.cell(row=row_idx, column=col_idx_charges).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE; ws_summary.cell(row=row_idx, column=col_idx_expenses).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE; ws_summary.cell(row=row_idx, column=col_idx_remaining).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

        # --- 8. Reorder Sheets ---
        print("[Step 8] Reordering sheets...")
        # Define desired order of main sheets
        desired_order_main = [SUMMARY_SHEET_NAME, SUMMARY_INDIVIDUAL_SHEET_NAME, NEEDS_REVIEW_SHEET_NAME]
        # Get club sheet names, sort them alphabetically
        club_sheet_titles_sorted = sorted([sheet.title for sheet in workbook.worksheets if sheet.title not in desired_order_main])
        # Combine main sheets and sorted club sheets for the final order
        desired_order_final = desired_order_main + club_sheet_titles_sorted
        # Reorder sheets in the workbook object
        sheets_by_title = {sheet.title: sheet for sheet in workbook._sheets} # Map titles to sheet objects
        workbook._sheets = [sheets_by_title[title] for title in desired_order_final if title in sheets_by_title] # Rebuild _sheets list
        print(f"  New sheet order: {[s.title for s in workbook._sheets]}")

        # --- [AUTO-FIT COLUMN WIDTHS - START] ---
        print("[Step 9] Adjusting column widths for readability...")
        for sheet_name in workbook.sheetnames: # Iterate through all sheets in the workbook
            if sheet_name in workbook: # Check if sheet exists (should always be true here)
                ws = workbook[sheet_name] # Get worksheet object
                column_widths = {} # Dictionary to store max width for each column
                # Iterate through all rows and cells to find the max width
                for row in ws.iter_rows(): 
                    for cell in row:
                        if cell.value: # If cell has a value
                            # Calculate length of cell value, with special handling for dates/numbers
                            if isinstance(cell.value, datetime): # For datetime objects
                                cell_len = 12 # Fixed width for dates like MM/DD/YYYY or YYYY-MM-DD
                            elif isinstance(cell.value, (int, float)): # For numbers
                                cell_len = len(f"{cell.value:.2f}") + 1 # Length of formatted number + sign
                            else: # For other types (strings)
                                cell_len = len(str(cell.value))
                            current_width = cell_len + 2 # Add padding
                            # Update max width for the column if current is greater
                            if cell.column_letter not in column_widths or current_width > column_widths[cell.column_letter]:
                                column_widths[cell.column_letter] = current_width
                # Apply the calculated widths to the columns
                for col_letter, width in column_widths.items():
                    adjusted_width = max(width, 8.43) # Ensure a minimum width (Excel default)
                    # Optional: Cap maximum width
                    # max_col_width = 60 
                    # ws.column_dimensions[col_letter].width = min(adjusted_width, max_col_width)
                    ws.column_dimensions[col_letter].width = adjusted_width
        # --- [AUTO-FIT COLUMN WIDTHS - END] ---

        # --- 9. Save Updated Workbook to Buffer --- (Now Step 10 due to new step for column widths)
        print("[Step 10] Saving updated workbook to memory buffer...") # Renumbered step
        output_buffer = BytesIO() # Create a BytesIO buffer
        workbook.save(output_buffer) # Save workbook to buffer
        output_buffer.seek(0) # Reset buffer position to the beginning
        print("  Workbook saved to buffer.")
        results['new_transactions_for_db'] = new_transactions_for_db # Add transactions for DB to results
    except FileNotFoundError as fnf_err: # Handle file not found errors
         err_msg = f"File not found: {fnf_err}"; print(f"[Error] {err_msg}"); results['errors'].append(err_msg)
    except ValueError as val_err: # Handle data validation errors (e.g., missing headers)
         err_msg = f"Data error: {val_err}"; print(f"[Error] {err_msg}"); results['errors'].append(err_msg)
    except Exception as e: # Catch any other unexpected errors
        err_msg = f"Unexpected error: {e}"; print(f"[Error] {err_msg}"); traceback.print_exc(); results['errors'].append(err_msg)
        if workbook: # Try to close workbook if it was opened
            try: workbook.close()
            except: pass # Ignore errors on close
            
    print(f"Update process finished. Results: {results}")
    print("="*20 + " Ending Foundation Summary Update " + "="*20 + "\n")
    return output_buffer, results
